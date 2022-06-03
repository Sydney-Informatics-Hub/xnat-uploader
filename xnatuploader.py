#!/usr/bin/env python


import argparse
import json

from pathlib import Path
from openpyxl import Workbook, load_workbook
from matcher import Matcher, XNAT_HIERARCHY
import xnatutils


def scan(matcher, root, logfile):
    """
    Scan the filesystem under root for files which match recipes and write
    out the resulting values to a spreadsheet.
    ---
    matcher: a PathMatcher
    root: pathlib.Path
    logfile: pathlib.Path
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["Recipe", "File", "Upload", "Status"] + XNAT_HIERARCHY + matcher.params)
    for file in root.glob("**/*"):
        filematch = matcher.match(file)
        ws.append(filematch.columns)
    wb.save(logfile)


def upload(xnat_session, matcher, project, logfile):
    """
    Load an Excel logfile created with scan and upload the files which the user
    has marked for upload, and which haven't been uploaded yet. Keeps track of
    successful uploads in the "success" column
    """
    wb = load_workbook(logfile)
    ws = wb.active
    header = True
    files = []
    for row in ws.values:
        if header:
            columns = row
            header = False
        else:
            matchfile = matcher.from_spreadsheet(row)
            files.append(matchfile)
    for file in files:
        if file.selected:
            if file.status == "success":
                print(f"{file.file} - already uploaded")
            else:
                try:
                    upload_file(xnat_session, project, file)
                    file.status = "success"
                except Exception as e:
                    file.error = str(e)
                    file.status = "failed"
    wb = Workbook()
    ws = wb.active
    ws.append(columns)
    for file in files:
        ws.append(file.columns)
    wb.save(logfile)


def upload_file(xnat_session, project, matchfile):
    xnatutils.put(
        matchfile.session,
        matchfile.dataset,
        matchfile.file,
        resource_name="DICOM",
        project_id=project,
        subject_id=matchfile.subject,
        create_session=True,
        connection=xnat_session,
    )


def show_help():
    pass


def main():
    ap = argparse.ArgumentParser("XNAT batch uploader")
    ap.add_argument("--config", default="config.json", type=Path)
    ap.add_argument("--source", default="./", type=Path)
    ap.add_argument("--log", default="log.xlsx", type=Path)
    ap.add_argument("--server", type=str)
    ap.add_argument("--project", type=str)
    ap.add_argument("operation", default="scan", choices=["scan", "upload", "help"])
    args = ap.parse_args()

    if args.operation == "help":
        show_help()
        exit()

    with open(args.config, "r") as fh:
        config_json = json.load(fh)
        matcher = Matcher(config_json)

    if args.operation == "scan":
        scan(matcher, args.source, args.log)
    else:
        if not args.server:
            print("Can't upload without a server")
            exit()
        if not args.project:
            print("Can't upload without a project ID")
            exit()
        xnat_session = xnatutils.base.connect(args.server)
        upload(xnat_session, matcher, args.project, args.log)


if __name__ == "__main__":
    main()
