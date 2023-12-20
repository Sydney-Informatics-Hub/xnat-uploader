#!/usr/bin/env python

import argparse
import csv
import logging
from tqdm import tqdm
from pathlib import Path
import xnatutils
import click
import re

from importlib.metadata import version

__version__ = version("xnatuploader")

from openpyxl import load_workbook

from xnatuploader.matcher import Matcher
from xnatuploader.dicoms import dicom_extractor, XNATFileMatch, SPREADSHEET_FIELDS
from xnatuploader.workbook import new_workbook, add_filesheet, load_config
from xnatuploader.upload import Upload, trigger_pipelines, parse_allow_fields

from xnatutils.base import sanitize_re

FILE_COLUMN_WIDTH = 50

DEBUG_MAX = 10

IGNORE_FILES = [".DS_Store"]

KEYBOARD_QUIT_STATUS = "Upload interrupted by user"
CONFIRM_KEYBOARD_QUIT_MSG = "Are you sure that you want to quit uploading?"

# re which matches errors that indicate that the project doesn't exist on
# XNAT or permissions are wrong: if this is encountered, the whole upload
# should be abandoned rather than repeatedly trying for each set of files

CANNOT_CREATE_RE = re.compile("Cannot create session")

logger = logging.getLogger(__name__)


def scan(
    matcher,
    root,
    spreadsheet,
    include_unmatched=True,
    strict_scan_ids=False,
    debug=False,
):
    """
    Scan the filesystem under root for files which match recipes and write
    out the resulting values to a new worksheet in the spreadsheet.

    If the debug flag is true, only try to match DEBUG_MAX files
    ---
    matcher: a Matcher
    root: pathlib.Path
    spreadsheet: pathlib.Path
    include_unmatched: boolean
    debug: boolean
    """
    logger.info(f"Loading {spreadsheet}")
    wb = load_workbook(spreadsheet)
    ws = add_filesheet(wb, matcher, debug)  # keeps old sheets if debug=True
    logger.info("Preparing file list")
    filepaths = sorted(
        [f for f in root.glob("**/*") if f.is_file() and f.name not in IGNORE_FILES]
    )
    files = []
    unmatched = []
    if debug:
        filepaths = filepaths[:DEBUG_MAX]
    logger.info(f"Scanning directory {root}")
    for filepath in tqdm(filepaths):
        if filepath.is_file():
            logger.debug(f"Scanning {filepath}")
            file = matcher.match(root, filepath)
            if file.success:
                logger.debug(f"Matched {file.file}")
                files.append(file)
            else:
                if include_unmatched:
                    file.load_dicom(matcher.extractor_options)
                    unmatched.append(file)

    skips, uploads = collate_uploads(files, strict_scan_ids)

    ns = len(uploads)
    nm = len(files)
    num = len(unmatched)

    if include_unmatched:
        logger.info(
            f"Saving {ns} scans with {nm} matching files and {num} non-matching files to {spreadsheet}"
        )
    else:
        logger.info(f"Saving {ns} scans with {nm} matching files to {spreadsheet}")

    for session_scan, upload in tqdm(uploads.items(), desc="Scans"):
        for file in upload.files:
            ws.append(file.columns)

    if include_unmatched:
        for file in unmatched:
            ws.append(file.columns)

    wb.save(spreadsheet)


def upload(
    xnat_session,
    matcher,
    project,
    spreadsheet,
    anon_rules=None,
    anonymize_files=False,
    strict_scan_ids=False,
    test=False,
    overwrite=False,
    no_pipeline=False,
):
    """
    Load an Excel spreadsheet created with scan and upload the files which the user
    has marked for upload, and which haven't been uploaded yet. Keeps track of
    successful uploads in the "status" column.

    Progress is written out to a temporary csv file. Files which are not being
    uploaded on this pass (because they were already uploaded, or because they
    weren't selected for upload) are still written out to the csv file.

    Exceptions during uploading are trapped and logged as failures in the
    spreadsheet, unless they're a KeyboardInterrupt. If one of these is
    recieved, the user is prompted to confirm that they want to stop, and then
    the files which haven't yet been uploaded are written out to the csv with
    a status message about the interrupt.
    ---
    xnat_session: an XnatPy session, as returned by xnatutils.base.connect
    matcher: a Matcher
    project: the XNAT project id to which we're uploading
    spreadsheet: pathlib.Path to the Excel spreadsheet listing files
    overwrite: Boolean, used to set the overwrite flag on xnatutils for testing
    """
    wb = load_workbook(spreadsheet)
    ws = wb["Files"]
    header = True
    files = []
    for row in ws.values:
        if header:
            header = False
        else:
            matchfile = matcher.from_spreadsheet(row)
            files.append(matchfile)
    skip, uploads = collate_uploads(files, strict_scan_ids)
    csvout = get_csv_filename(spreadsheet)
    if test:
        dry_run(uploads)
        return
    written = {}
    abandoned = False
    with open(csvout, "w", newline="") as cfh:
        csvw = csv.writer(cfh)
        for file in skip:
            csvw.writerow(file.columns)
        keyboard_quit = False
        for session_scan, upload in tqdm(uploads.items(), desc="Sessions"):
            logger.debug(f"Uploading {session_scan}")
            try:
                upload.start_upload(xnat_session, project)
                for file in tqdm(upload.files, desc=session_scan):
                    logger.debug(f"Uploading {file.file}")
                    try:
                        status = upload.upload(
                            [file],
                            anonymize_files=anonymize_files,
                            overwrite=overwrite,
                            anon_rules=anon_rules,
                        )
                        file.status = status[file.file]
                    except KeyboardInterrupt:
                        if click.confirm(CONFIRM_KEYBOARD_QUIT_MSG):
                            keyboard_quit = True
                            logger.warning(
                                f"KeyboardInterrupt in file loop {file.file}"
                            )
                            break
                    except Exception as e:
                        file.status = log_failure(f"File {file.file}", e)
                    csvw.writerow(file.columns)
                    written[file.file] = True
            except KeyboardInterrupt:
                if click.confirm(CONFIRM_KEYBOARD_QUIT_MSG):
                    keyboard_quit = True
                    logger.warning("KeyboardInterrupt in dataset loop")
                    break
            except Exception as e:
                if CANNOT_CREATE_RE.match(str(e)):
                    log_failure(f"Dataset {upload.label}", e)
                    logger.error(
                        f"Check that project {project} exists and "
                        "you have upload permissions"
                    )
                    abandoned = True
                    break
                status = log_failure(f"Dataset {upload.label}", e)
                for file in upload.files:
                    file.status = status
                    csvw.writerow(file.columns)
                    written[file.file] = True
            if keyboard_quit:
                break
        if not no_pipeline:
            trigger_pipelines(xnat_session, project, uploads)

        if keyboard_quit:
            for _, upload in tqdm(uploads.items(), desc="Updating spreadsheet"):
                for file in upload.files:
                    if file.file not in written:
                        file.status = KEYBOARD_QUIT_STATUS
                        csvw.writerow(file.columns)
    if not abandoned:
        copy_csv_to_spreadsheet(matcher, csvout, spreadsheet)


def log_failure(label, e):
    """Write a message about a file or dataset upload failure to the logs,
    and return a value to be recorded in the spreadsheet. It's in its own
    function to make the loop in upload a bit easier to read.
    """
    error = str(e)
    logger.error(f"{label} exception: {error}")
    return error


def dry_run(uploads):
    """Just logs what would be uploaded - broken out of the main upload
    loop because it doesn't need to do anything else
    ---
    uploads: list of Upload
    """
    for session_label, upload in uploads.items():
        logger.debug(f"Uploading {session_label}")
        upload.log(logger)


def copy_csv_to_spreadsheet(matcher, csvout, spreadsheet):
    """Copies the csv of uploaded files to the Files worksheet of the
    spreadsheet. If it can't, tells the user that the results are in the csv
    file.
    This function always clobbers the Files worksheet with its updated value,
    unlike scan, which can save old versions of Files when running in debug
    mode.
    """
    wb = load_workbook(spreadsheet)
    ws = add_filesheet(wb, matcher, False)
    logger.debug(f"Copying upload results from {csvout} to {spreadsheet}")
    with open(csvout, "r") as cfh:
        for row in csv.reader(cfh):
            ws.append(row)
    try:
        wb.save(spreadsheet)
    except PermissionError:
        logger.error(
            f"""
A permissions error prevented the script from writing the upload results back
to {spreadsheet}.  If you are on Windows, this may be because you still have
the spreadsheet open in Excel.

The results are available as a CSV file: {csvout}
"""
        )


def collate_uploads(files, strict_scan_ids):
    """
    Takes a list of files and collates them by subject (patient), visit
    index (starting from the earliest), scan type, and (optionally) scan_id,
    returning a list of files which have skipped or already uploaded and a dictionary
    of Uploads keyed by {session_label}_{scan_id}

    ---
    files: list of FileMatch

    returns: tuple of ( list of FileMatch, dict of str: Upload )
    """

    subjects = {}
    skip = []
    for file in files:
        if not file.selected:
            skip.append(file)
        else:
            if file.status == "success":
                logger.debug(f"skipping file already uploaded {file.file}")
                skip.append(file)
            else:
                if file["Subject"] not in subjects:
                    subjects[file["Subject"]] = []
                subjects[file["Subject"]].append(file)
    uploads = {}
    for subject_id, files in subjects.items():
        dates = sorted(set([file.study_date for file in files]))
        visits = {dates[i]: i + 1 for i in range(len(dates))}
        clean_datasets = sanitise_dataset_names(files)
        for file in files:
            visit = visits[file.study_date]
            modality = file.modality
            scan_id = file.series_number
            if strict_scan_ids:
                session_label = f"{subject_id}_{modality}{visit}_{scan_id}"
            else:
                session_label = f"{subject_id}_{modality}{visit}"
            file.session_label = session_label
            scan_type = clean_datasets[file.dataset]
            session_scan = f"{session_label}:{scan_type}"
            if session_scan not in uploads:
                uploads[session_scan] = Upload(
                    session_label=session_label,
                    subject=subject_id,
                    date=file.study_date,
                    modality=modality,
                    series_number=scan_id,
                    scan_type=scan_type,
                    strict_scan_ids=strict_scan_ids,
                    manufacturer=file.manufacturer,
                    model=file.model,
                )
            uploads[session_scan].add_file(file)
    return skip, uploads


def sanitise_dataset_names(files):
    """
    For a list of files, sanitise the .dataset values (replace characters which
    XNAT doesn't allow in a resource id with '_') and then make sure that the
    datasets are all still unique by appending 1, 2, etc to them.

    Returns a dict which maps the original datasets to sanitised values
    ---
    files: list of FileMatch

    returns: dict of str: str
    """
    clean = {}
    used = []
    for file in files:
        if file.dataset not in clean:
            base = sanitize_re.sub("_", file.dataset)
            sanitised = base
            i = 1
            while sanitised in used:
                i += 1
                sanitised = base + str(i)
            clean[file.dataset] = sanitised
            used.append(sanitised)
    return clean


def get_csv_filename(spreadsheet):
    csv = spreadsheet.with_suffix(".csv")
    n = 0
    while csv.is_file():
        n += 1
        csv = spreadsheet.parent / Path(f"{spreadsheet.stem}.{n}.csv")
    return csv


def show_help():
    print(
        """
xnatuploader is a utility for collating and uploading images to XNAT, using
a spreadsheet to keep track of which files are uploaded.

Sample usage:

    xnatuploader --spreadsheet sheet.xlsx init

Writes out a spreadsheet in the format required by xnatuploader, with a sample
configuration page.

    xnatuploader --spreadsheet sheet.xlsx --dir ./source/ scan

Scans the directory provided with the --dir flag and builds a file list in
the spreadsheet

    xnatuploader --spreadsheet sheet.xlsx --dir ./source upload

Uploads the files recorded in the spreadsheet, to the server and project
specified in the config worksheet.

For more detailed instructions on how to configure xnatuploader to capture
parameters from filepaths, refer to the "Configuration" worksheet in the
spreadsheet, or visit the online documentation at:

https://github.com/Sydney-Informatics-Hub/xnat-uploader/

"""
    )


def opt_or_config(args, config, param):
    """
    Gets a config value from either a command line parameter or the config
    spreadsheet (in that order). Raises a ValueException if no value is
    available.
    ---
    args: Namespace as returned by argparse
    config: the config dict
    value: the key in the config dict
    """
    value = None
    params = vars(args)
    if param.lower() in params:
        value = params[param.lower()]
    if value is None:
        if param in config:
            value = config[param]
    if value is None:
        raise ValueError(f"{param} must be specified in config or via command line")
    return value


def main():
    ap = argparse.ArgumentParser("XNAT batch uploader")
    ap.add_argument(
        "--dir", default="./", type=Path, help="Base directory to scan for files"
    )
    ap.add_argument(
        "--spreadsheet", default="files.xlsx", type=Path, help="File list spreadsheet"
    )
    ap.add_argument("--server", type=str, help="XNAT server")
    ap.add_argument("--project", type=str, help="XNAT project ID")
    ap.add_argument("--loglevel", type=str, default="info", help="Logging level")
    ap.add_argument(
        "--debug",
        action="store_true",
        default=False,
        help=f"""
Debug mode: only attempt to match {DEBUG_MAX} patterns and generates a lot of
debug messages
""",
    )
    ap.add_argument(
        "--logdir", type=Path, default="logs", help="Directory to write logs to"
    )
    ap.add_argument(
        "--test",
        action="store_true",
        default=False,
        help="Test mode: don't upload, just log what will be uploaded",
    )
    ap.add_argument(
        "--unmatched",
        action="store_true",
        default=False,
        help="Whether to include unmatched files in list",
    )
    ap.add_argument(
        "--strict",
        action="store_true",
        default=False,
        help="Whether to collate uploads by series number / scan id",
    )
    ap.add_argument(
        "--anonymize",
        action="store_true",
        default=False,
        help="Whether to anonymize files before uploading",
    )
    ap.add_argument(
        "--overwrite",
        action="store_true",
        default=False,
        help="Whether to overwrite files which have already been uploaded",
    )
    ap.add_argument(
        "--nopipeline",
        action="store_true",
        default=False,
        help="Don't trigger the metadata extraction and pipeline",
    )
    ap.add_argument("--version", action="version", version="%(prog)s " + __version__)
    ap.add_argument(
        "operation",
        default="scan",
        choices=["init", "scan", "upload", "help"],
        help="Operation",
    )
    args = ap.parse_args()

    if not args.logdir.is_dir():
        args.logdir.mkdir(parents=True)

    loglevel = args.loglevel.upper()
    if args.debug:
        loglevel = "DEBUG"

    logger.setLevel(logging.DEBUG)
    logfh = logging.FileHandler(args.logdir / "xnatuploader.log")
    logfh.setLevel(args.loglevel.upper())
    logger.addHandler(logfh)
    logch = logging.StreamHandler()
    logch.setLevel(args.loglevel.upper())
    logger.addHandler(logch)

    if args.operation == "help":
        show_help()
        exit()

    if args.operation == "init":
        new_workbook(args.spreadsheet)
        logger.info(f"Initialised spreadsheet at {args.spreadsheet}")
        exit()

    config = load_config(args.spreadsheet)
    skip_image_types = []
    if "SkipImageTypes" in config["xnat"]:
        skip_image_types = config["xnat"]["SkipImageTypes"]
        if skip_image_types is None:
            skip_image_types = []
        else:
            skip_image_types = skip_image_types.split(",")

    matcher = Matcher(
        patterns=config["paths"],
        mappings=config["mappings"],
        fields=SPREADSHEET_FIELDS,
        file_extractor=dicom_extractor,
        extractor_options={"skip_image_types": skip_image_types},
        match_class=XNATFileMatch,
        loglevel=loglevel,
    )

    if args.operation == "scan":
        scan(
            matcher,
            args.dir,
            args.spreadsheet,
            include_unmatched=args.unmatched,
            strict_scan_ids=args.strict,
            debug=args.debug,
        )
    else:
        server = opt_or_config(args, config["xnat"], "Server")
        project = opt_or_config(args, config["xnat"], "Project")
        anon_rules = opt_or_config(args, config["xnat"], "AllowFields")
        # not using opt_or_config for AllowFields as it's config only
        anon_rules = {}
        if "AllowFields" in config["xnat"]:
            anon_rules = parse_allow_fields(config["xnat"]["AllowFields"])
        logger.debug(f"Server = {server}")
        logger.debug(f"Project = {project}")
        xnat_session = xnatutils.base.connect(server)
        logger.debug(f"main anon rules {anon_rules}")
        upload(
            xnat_session,
            matcher,
            project,
            args.spreadsheet,
            strict_scan_ids=args.strict,
            anonymize_files=args.anonymize,
            anon_rules=anon_rules,
            test=args.test,
            overwrite=args.overwrite,
            no_pipeline=args.nopipeline,
        )


if __name__ == "__main__":
    main()
