from collections import OrderedDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles.alignment import Alignment

FILE_COLUMN_WIDTH = 50
HELP_COLUMN_WIDTH = 25
HELP_ROW_HEIGHT = 90

HELP_TEXT = """
Paths are a set of patterns to be matched against file paths in the source directory.

Values in curly brackets like {SubjectName} are captured into variables.
Variables with names in all caps like {YYYY} will only match numbers.
The special value * and ** match one or more than one intervening directory.

Mappings are used to combine the variables captured from a path to the
XNAT hierarchy Subject / Session / Dataset.

You can also use metadata from the DICOM files in the mappings, by
using the syntax "DICOM:MetadataName" - for example, DICOM:StudyDate

The XNAT section configures the XNAT project ID and server.
"""


class WorkbookError(Exception):
    pass


def new_workbook(file):
    """
    Make a new spreadsheet with the instructions worksheet and default
    patterns
    --
    file: filepath.Path
    """
    wb = Workbook()
    ws = wb.active
    for col in "ABCDE":
        ws.column_dimensions[col].width = HELP_COLUMN_WIDTH

    ws.title = "Configuration"
    ws["A1"] = "Instructions"
    ws["A2"] = HELP_TEXT
    ws["A2"].alignment = Alignment(wrapText=True)
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = HELP_ROW_HEIGHT
    ws["A4"] = "Configuration"
    ws["A5"] = "Paths"
    ws["B5"] = "DICOM"
    ws["C5"] = "{SubjectName}-{ID}"
    ws["D5"] = "**"
    ws["E5"] = "{Directory}"
    ws["F5"] = "{filename}.dcm"
    ws["A10"] = "Mappings"
    ws["B10"] = "Subject"
    ws["C10"] = "ID"
    ws["B11"] = "Session"
    ws["C11"] = "DICOM:StudyDate"
    ws["B12"] = "Dataset"
    ws["C12"] = "Directory"
    ws["A14"] = "XNAT"
    ws["B14"] = "Project"
    ws["C14"] = "Test001"
    ws["B15"] = "Server"
    ws["C15"] = "http://localhost:8080"
    ws["B16"] = "AllowFields"
    ws["C16"] = "AccessionNumber"
    ws["B17"] = "SkipImageTypes"
    ws["C17"] = "DOSE_INFO"
    wb.save(file)


def add_filesheet(wb, matcher, debug):
    """
    Adds a worksheet to a workbook with headers from the matcher object
    and niceties such as a wider column for the file names.

    If debug is true, an existing worksheet with the name "Files" is renamed
    to "Files-prev" - otherwise it's deleted.
    ---
    wb: a Workbook
    matcher: a Matcher
    debug: bool

    returns: the new worksheet
    """
    if "Files" in wb:
        if debug:
            old_files = wb["Files"]
            # older versions will be automatically renamed to Files-prev1,
            # Files-prev2 etc
            old_files.title = "Files-prev"
        else:
            wb.remove(wb["Files"])
    ws = wb.create_sheet("Files")
    ws.column_dimensions["B"].width = FILE_COLUMN_WIDTH
    ws.append(matcher.headers)
    return ws


def load_config(excelfile):
    """
    Load config from the Configuration worksheet of the spreadsheet.
    Each section is loaded into an OrderedDict - this is because order is
    significant for building the columns in the spreadsheet.
    """
    wb = load_workbook(excelfile)
    if "Configuration" not in wb:
        raise WorkbookError(f"No worksheet named 'Configuration' in {excelfile}")
    ws = wb["Configuration"]
    config = {}
    section = None
    sections = ["paths", "mappings", "xnat"]
    for row in ws:
        if row[0].value is not None:
            section = row[0].value.lower()
        var = row[1].value
        if var is not None and section is not None:
            if section not in config:
                config[section] = OrderedDict()
            cells = [cell.value for cell in row[2:] if cell.value is not None]
            if section == "xnat":
                try:
                    config[section][var] = cells[0]
                except IndexError:
                    config[section][var] = None
            else:
                config[section][var] = cells
    missing = [s for s in sections if s not in config]
    if len(missing) > 0:
        raise WorkbookError(f"Missing config sections: {missing}")
    return config
