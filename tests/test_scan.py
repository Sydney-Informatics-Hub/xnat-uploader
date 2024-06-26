import logging
import json
from openpyxl import load_workbook
from pathlib import Path
import pytest

from xnatuploader.matcher import Matcher
from xnatuploader.dicoms import dicom_extractor, XNATFileMatch, SPREADSHEET_FIELDS
from xnatuploader.xnatuploader import scan, collate_uploads
from xnatuploader.workbook import load_config, new_workbook

logger = logging.getLogger(__name__)


def assert_worksheets_equal(expect, got):
    grows = list(got.values)
    for row in expect.values:
        assert grows[0] == row
        grows = grows[1:]
    assert len(grows) == 0


def test_scan(tmp_path, test_files):
    fileset = test_files["basic"]
    config = load_config(fileset["config_excel"])
    matcher = Matcher(
        config["paths"],
        config["mappings"],
        SPREADSHEET_FIELDS,
        dicom_extractor,
        XNATFileMatch,
    )
    scanned = tmp_path / "scanned.xlsx"
    new_workbook(scanned)
    scan(matcher, Path(fileset["dir"]), scanned, include_unmatched=True)
    expect_wb = load_workbook(fileset["scanned_excel"])
    got_wb = load_workbook(scanned)
    assert "Files" in got_wb
    assert_worksheets_equal(expect_wb["Files"], got_wb["Files"])


@pytest.mark.parametrize("source_dir", ["basic", "basic_strict"])
def test_collation(source_dir, tmp_path, test_files):
    fileset = test_files[source_dir]
    uploads_dict = fileset["uploads_dict"]
    config = load_config(fileset["config_excel"])
    matcher = Matcher(
        config["paths"],
        config["mappings"],
        SPREADSHEET_FIELDS,
        dicom_extractor,
        XNATFileMatch,
    )
    log = tmp_path / "log.xlsx"
    new_workbook(log)
    scan(matcher, Path(fileset["dir"]), log, strict_scan_ids=fileset["strict_scan_ids"])
    wb = load_workbook(log)
    ws = wb["Files"]
    header = True
    files = []
    for row in ws.values:
        if header:
            header = False
        else:
            matchfile = matcher.from_spreadsheet(row)
            files.append(matchfile)
    skipped, uploads = collate_uploads(files, fileset["strict_scan_ids"])
    for session_scan, upload in uploads.items():
        uploads[session_scan] = [f.file for f in uploads[session_scan].files]
    assert uploads == uploads_dict["uploads"]
    assert len(skipped) == uploads_dict["skipped"]


def test_sanitisation_collisions(tmp_path, test_files, sanitised_dict):
    fileset = test_files["sanitisation"]
    config_file = fileset["config"]
    with open(config_file, "r") as fh:
        config = json.load(fh)
    matcher = Matcher(
        config["paths"],
        config["mappings"],
        SPREADSHEET_FIELDS,
        dicom_extractor,
        XNATFileMatch,
    )
    logger.warning(f"Testing santisation in {fileset}")
    log = tmp_path / "log.xlsx"
    new_workbook(log)
    scan(matcher, Path(fileset["dir"]), log, strict_scan_ids=True)
    wb = load_workbook(log)
    ws = wb["Files"]
    header = True
    files = []
    for row in ws.values:
        if header:
            header = False
        else:
            matchfile = matcher.from_spreadsheet(row)
            files.append(matchfile)
    logger.warning(f"File list = {files}")
    skipped, uploads = collate_uploads(files, True)
    for session_scan, upload in uploads.items():
        uploads[session_scan] = [f.file for f in uploads[session_scan].files]
    assert uploads == sanitised_dict["uploads"]
    assert len(skipped) == sanitised_dict["skipped"]


@pytest.mark.parametrize("source_dir", ["basic", "basic_strict"])
def test_collation_skips(tmp_path, test_files, source_dir):
    fileset = test_files[source_dir]
    uploads_dict = fileset["uploads_dict"]
    config = load_config(fileset["config_excel"])
    matcher = Matcher(
        config["paths"],
        config["mappings"],
        SPREADSHEET_FIELDS,
        dicom_extractor,
        XNATFileMatch,
    )
    log = tmp_path / "log.xlsx"
    new_workbook(log)
    scan(matcher, Path(fileset["dir"]), log, strict_scan_ids=fileset["strict_scan_ids"])
    wb = load_workbook(log)
    ws = wb["Files"]
    header = True
    files = []
    skip_n = 4  # mark the first four files as "success"
    n = 0
    for row in ws.values:
        if header:
            header = False
        else:
            matchfile = matcher.from_spreadsheet(row)
            if n < skip_n:
                if matchfile.selected:
                    matchfile.status = "success"
                    logger.warning(f"Skipping {matchfile.file}")
                    n += 1
            files.append(matchfile)
    skip, uploads = collate_uploads(files, fileset["strict_scan_ids"])
    assert len(skip) == uploads_dict["skipped"] + skip_n
    for session_scan, upload in uploads.items():
        uploads[session_scan] = [f.file for f in uploads[session_scan].files]
    # remove skipped files from the fixture dictionary
    upload_skipped = {}
    skipstr = [f.file for f in skip]
    for subject_id, files in uploads_dict["uploads"].items():
        upload_files = [f for f in files if f not in skipstr]
        # don't add if the list of files is empty
        if upload_files:
            upload_skipped[subject_id] = upload_files
    assert uploads == upload_skipped


@pytest.mark.parametrize("forbidden", ["secret_pdf", "dose_info"])
def test_secret_pdfs(tmp_path, test_files, forbidden):
    fileset = test_files[forbidden]
    config = load_config(fileset["config_excel"])
    matcher = Matcher(
        config["paths"],
        config["mappings"],
        SPREADSHEET_FIELDS,
        dicom_extractor,
        XNATFileMatch,
    )
    scanned = tmp_path / "scanned.xlsx"
    new_workbook(scanned)
    scan(
        matcher,
        Path(fileset["dir"]),
        scanned,
        include_unmatched=True,
        strict_scan_ids=True,
    )
    expect_wb = load_workbook(fileset["scanned_excel"])
    got_wb = load_workbook(scanned)
    assert "Files" in got_wb
    assert_worksheets_equal(expect_wb["Files"], got_wb["Files"])
