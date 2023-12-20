import xnatutils
import shutil
import json
import pytest

from openpyxl import load_workbook

from pathlib import Path

from xnatuploader.xnatuploader import scan, upload
from xnatuploader.matcher import Matcher
from xnatuploader.dicoms import dicom_extractor, XNATFileMatch, SPREADSHEET_FIELDS
from xnatuploader.workbook import new_workbook
from xnatuploader.upload import parse_allow_fields


@pytest.mark.parametrize("anonymize", [False, True])
@pytest.mark.parametrize("source_dir", ["basic", "bad_paths"])
def test_upload_from_spreadsheet(
    source_dir, anonymize, xnat_connection, tmp_path, test_files
):
    test_config = test_files[source_dir]["config"]
    test_dir = test_files[source_dir]["dir"]
    with open(test_config, "r") as fh:
        config = json.load(fh)
        matcher = Matcher(
            config["paths"],
            config["mappings"],
            SPREADSHEET_FIELDS,
            dicom_extractor,
            {"skip_image_types": []},
            XNATFileMatch,
        )
        anon_rules = parse_allow_fields(config["xnat"]["AllowFields"])
    project = xnat_connection.classes.ProjectData(
        parent=xnat_connection,
        name="Test_" + source_dir,
    )
    log_scanned = tmp_path / "log_scanned.xlsx"
    log_uploaded = tmp_path / "log_uploaded.xlsx"
    downloads = tmp_path / "downloads"
    new_workbook(log_scanned)
    scan(matcher, Path(test_dir), log_scanned)
    shutil.copy(log_scanned, log_uploaded)
    upload(
        xnat_connection,
        matcher,
        project.name,
        log_uploaded,
        strict_scan_ids=False,
        anonymize_files=anonymize,
        anon_rules=anon_rules,
        overwrite=True,
    )
    uploaded_wb = load_workbook(log_uploaded)
    uploaded_ws = uploaded_wb["Files"]
    uploads = {}
    for row in uploaded_ws.values:
        if row[0] != "Recipe":
            upload_row = matcher.from_spreadsheet(row)
            if upload_row.selected:
                uploads[upload_row.file] = upload_row
    expect = {}
    for row in uploaded_ws.values:
        if row[0] != "Recipe":
            m = matcher.from_spreadsheet(row)
            if m.selected:
                subject = m.subject
                session_label = m.session_label
                assert session_label is not None
                if subject not in expect:
                    expect[subject] = {}
                if session_label not in expect[subject]:
                    expect[subject][session_label] = []
                expect[subject][session_label].append(m)
    for subject, sessions in expect.items():
        assert subject in project.subjects
        for session_label, rows in sessions.items():
            assert session_label in project.experiments
            xnat_date = project.experiments[session_label].date
            assert xnat_date.strftime("%Y%m%d") == rows[0].study_date
            for row in rows:
                assert row.file in uploads
                assert uploads[row.file].status == "success"
            # Note: passing the subject_id in is required to get this to not
            # break on the second set of test cases if they are being done in
            # the same project
            xnatutils.get(
                session_label,
                downloads,
                project_id=project.name,
                subject_id=subject,
                connection=xnat_connection,
            )
            if session_label is not None:
                downloaded = get_downloaded(downloads / session_label)
                assert len(downloaded) == len(rows)
                for row in rows:
                    assert Path(row.file).name in downloaded
                    del uploads[row.file]
    assert len(uploads) == 0


def get_downloaded(session_download):
    files = []
    for child in session_download.iterdir():
        if child.is_dir():
            for file in child.iterdir():
                if file.is_file() and file.suffix == ".dcm":
                    files.append(file.name)
    return files


def test_missing_file(xnat_connection, tmp_path, test_files):
    project = xnat_connection.classes.ProjectData(
        parent=xnat_connection,
        name="Test_missing",
    )
    basic = test_files["basic"]
    with open(basic["config"], "r") as fh:
        config = json.load(fh)
        matcher = Matcher(
            config["paths"],
            config["mappings"],
            SPREADSHEET_FIELDS,
            dicom_extractor,
            {"skip_image_types": []},
            XNATFileMatch,
        )
        anon_rules = parse_allow_fields(config["xnat"]["AllowFields"])
    log_scanned = tmp_path / "log_scanned.xlsx"
    log_uploaded = tmp_path / "log_uploaded.xlsx"
    new_workbook(log_scanned)
    scan(matcher, Path(basic["dir"]), log_scanned)
    scanned_wb = load_workbook(log_scanned)
    ws = scanned_wb["Files"]
    # change the third filename to trigger an error
    c = 0
    i = 0
    bad_file = None
    for row in ws.values:
        i += 1
        if row[0] != "Recipe":
            m = matcher.from_spreadsheet(row)
            if m.selected:
                c += 1
                if c == 3:
                    bad_file = ws.cell(i, 2).value + "broken"
                    ws.cell(i, 2).value = bad_file

    scanned_wb.save(log_uploaded)
    upload(
        xnat_connection,
        matcher,
        project.name,
        log_uploaded,
        anonymize_files=False,
        anon_rules=anon_rules,
        strict_scan_ids=False,
        overwrite=True,
    )
    uploads = {}
    uploaded_wb = load_workbook(log_uploaded)
    uploaded_ws = uploaded_wb["Files"]
    for row in uploaded_ws.values:
        if row[0] != "Recipe":
            upload_row = matcher.from_spreadsheet(row)
            if upload_row.selected:
                uploads[upload_row.file] = upload_row
                if upload_row.file == bad_file:
                    assert upload_row.status != "success"
                else:
                    assert upload_row.status == "success"
